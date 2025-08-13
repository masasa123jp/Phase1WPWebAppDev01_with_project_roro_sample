#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
geocode_excel_nominatim_v3.py
=============================

Excel（.xlsx / .xlsm）の複数シートに対して、住所から緯度・経度（WGS84）を
無料のジオコーディングサービス（OpenStreetMap Nominatim）で取得し、
各シートに「緯度」「経度」列を付与して保存します。

【本版の改善点（v3）】
- .xlsm のマクロ保持保存で発生する「Workbook.__init__() got an unexpected keyword argument 'keep_vba'」
  問題に対し、openpyxl.load_workbook(keep_vba=True) を併用する実装に変更。
  → pandas/openpyxl のバージョン差異に依存せず安定してマクロを保持可能。
- 日本住所の正規化とフォールバック戦略を強化（表記ゆれ、括弧・建物・階等の削除、郵便番号の活用）。
- 成功結果のみキャッシュ保存（失敗はキャッシュしない）で、住所修正後の再試行を阻害しない。
- 既存の緯度/経度（別名列を含む）を尊重し、空欄だけ補完。

■ 使い方（例）
1) ライブラリ導入
   pip install pandas openpyxl requests tqdm python-dotenv

2) 実行例（.xlsm のマクロを保持して上書き/別名保存）
   python geocode_excel_nominatim_v3.py "【作業中】地図情報_マスタデータ.xlsm" \
       --sheets "シート1" "シート2" \
       --keep-vba \
       --output "【作業中】地図情報_マスタデータ_geocoded.xlsm"

   ※シート名や住所列名が不明な場合は省略可（自動検出）。

3) .env（推奨：問い合わせ元連絡先の明示）
   GEO_EMAIL=you@example.com

■ Nominatim 利用上の注意
- 最低 1 秒/リクエストのレート制限を遵守（本スクリプトで制御）
- 大量処理はローカルキャッシュを活用
- 高負荷/商用は有償サービスも検討

Author: ChatGPT (GPT-5 Thinking)
Date  : 2025-08-13 (Asia/Tokyo)
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
import time
import json
import re
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests
from requests.adapters import HTTPAdapter, Retry
from tqdm import tqdm
from openpyxl import load_workbook

# -------------------------------
# 設定
# -------------------------------

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
MIN_REQUEST_INTERVAL_SEC = 1.0  # Nominatim 推奨（最低1秒）

# 住所候補の既定カラム名（優先順）
DEFAULT_ADDRESS_CANDIDATE_COLS: List[str] = [
    "住所", "所在地", "Address", "address", "所在地住所", "住所1", "住所２", "住所2",
    "所在地1", "所在地2"
]

# 住所の構成要素候補（分割列がある場合の連結用）
DEFAULT_COMPONENT_COL_GROUPS: List[List[str]] = [
    ["都道府県", "pref", "prefecture", "都"],
    ["市区町村", "city", "区市町村", "市", "区", "町", "村"],
    ["町名", "番地", "丁目", "地番", "番", "号", "町域", "大字", "小字", "丁"],
    ["建物名", "ビル名", "マンション名", "建屋", "建物", "号室", "階"],
]

# 郵便番号候補カラム
DEFAULT_POSTAL_COLS: List[str] = [
    "郵便番号", "郵便", "郵便No", "Zip", "ZIP", "PostCode", "Postcode", "PostalCode", "postal_code"
]

# 出力列名
LAT_COL = "緯度"
LON_COL = "経度"


@dataclass
class GeocodeResult:
    lat: Optional[float]
    lon: Optional[float]
    source: str
    raw: Dict


# -------------------------------
# 正規化ユーティリティ
# -------------------------------

def normalize_space(s: str) -> str:
    """空白や全角空白を単一半角スペースに正規化。"""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u3000", " ")  # 全角空白→半角
    s = re.sub(r"\s+", " ", s.strip())
    return s


def to_halfwidth_basic(s: str) -> str:
    """数字・記号・スペースを中心に全角→半角へ簡易変換。"""
    if not s:
        return s
    try:
        import unicodedata
        s = unicodedata.normalize("NFKC", s)
    except Exception:
        pass
    return s


def unify_dashes(s: str) -> str:
    """ダッシュ類（ー—−–~～など）をハイフンに統一。"""
    if not s:
        return s
    return re.sub(r"[ー—−–~～﹘﹣‐‑‒―]", "-", s)


def strip_parentheses(s: str) -> str:
    """括弧と中身を削除。"""
    if not s:
        return s
    return re.sub(r"[（\(].*?[）\)]", "", s)


def strip_building_details(s: str) -> str:
    """建物/階/号室等の末尾表記を削除。"""
    if not s:
        return s
    # よくある語をカット（末尾付近）
    patterns = [
        r"\b\d+F\b", r"\bF\d+\b", r"\d+階", r"\d+号室", r"\bRoom\s*\d+\b", r"\b#\s*\d+\b",
        r"地下\d+階", r"地上\d+階", r"別館", r"本館"
    ]
    out = s
    for p in patterns:
        out = re.sub(p, "", out, flags=re.IGNORECASE)
    # 末尾の区切り記号を落とす
    out = re.sub(r"[ \-‐‑‒–—―、。,\.]+$", "", out)
    out = normalize_space(out)
    return out


def simplify_block_level(s: str) -> str:
    """丁目・番地・号などを簡略化（広域検索用）。"""
    if not s:
        return s
    s2 = re.sub(r"\d+(-\d+)*号?", "", s)  # 1-2-3 号 → 削る
    s2 = re.sub(r"\d+丁目", "", s2)
    s2 = re.sub(r"\d+番地?", "", s2)
    s2 = normalize_space(s2)
    return s2


def detect_postal_code(row: pd.Series, postal_cols: List[str]) -> Optional[str]:
    """行から郵便番号を検出して 123-4567 形式に正規化。"""
    for c in row.index:
        if c in postal_cols or c.lower() in [pc.lower() for pc in postal_cols]:
            val = str(row[c]) if pd.notna(row[c]) else ""
            val = to_halfwidth_basic(val)
            m = re.search(r"(\d{3})[-‐‑‒–—―]?\s?(\d{4})", val)
            if m:
                return f"{m.group(1)}-{m.group(2)}"
    return None


def build_address_from_components(row: pd.Series, component_groups: List[List[str]]) -> Optional[str]:
    """都道府県/市区町村/番地/建物等の分割列が存在する場合に連結して住所を構築。"""
    parts: List[str] = []
    for group in component_groups:
        chosen = None
        for col in row.index:
            if col in group or col.lower() in [g.lower() for g in group]:
                val = normalize_space(row[col])
                if val:
                    chosen = val
                    break
        if chosen:
            parts.append(chosen)
    if parts:
        return " ".join(parts)
    return None


def detect_address_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """候補名の優先順で住所列を検出（ケース無視の完全一致も考慮）。"""
    for name in candidates:
        if name in df.columns:
            return name
    lower_map = {c.lower(): c for c in df.columns}
    for name in candidates:
        if name.lower() in lower_map:
            return lower_map[name.lower()]
    return None


def prepare_address_series(df: pd.DataFrame,
                           address_col: Optional[str],
                           component_groups: List[List[str]]) -> pd.Series:
    """住所カラムがなければ構成要素から合成。どちらもなければ空を返す。"""
    if address_col and address_col in df.columns:
        return df[address_col].map(lambda x: normalize_space(to_halfwidth_basic(str(x)))).fillna("")
    # 分割列の連結を試みる
    addr_list: List[str] = []
    for _, row in df.iterrows():
        addr = build_address_from_components(row, component_groups) or ""
        addr_list.append(addr)
    return pd.Series(addr_list, index=df.index, name="__address_built__")


# -------------------------------
# Nominatim クライアント & キャッシュ
# -------------------------------

class NominatimClient:
    """Nominatim（OSM）向けの薄いクライアント。レート制御とリトライを内包。"""

    def __init__(self, email: Optional[str] = None, lang: str = "ja", countrycodes: str = "jp"):
        self.email = email or os.getenv("GEO_EMAIL")
        self.lang = lang
        self.countrycodes = countrycodes
        self._last_request_ts = 0.0

        self.session = requests.Session()
        retries = Retry(
            total=5,
            backoff_factor=1.2,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"]
        )
        self.session.mount("https://", HTTPAdapter(max_retries=retries))

    def _user_agent(self) -> str:
        base = "ProjectRoro-Geocoder/1.1"
        if self.email:
            return f"{base} contact:{self.email}"
        return base

    def _respect_rate_limit(self):
        now = time.time()
        elapsed = now - self._last_request_ts
        if elapsed < MIN_REQUEST_INTERVAL_SEC:
            time.sleep(MIN_REQUEST_INTERVAL_SEC - elapsed)

    def geocode_once(self, query: str) -> GeocodeResult:
        params = {
            "q": query,
            "format": "jsonv2",
            "addressdetails": 0,
            "limit": 1,
            "accept-language": self.lang,
            "countrycodes": self.countrycodes,
        }
        headers = {"User-Agent": self._user_agent()}

        self._respect_rate_limit()
        resp = self.session.get(NOMINATIM_URL, params=params, headers=headers, timeout=30)
        self._last_request_ts = time.time()
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list) and data:
            try:
                lat = float(data[0]["lat"])
                lon = float(data[0]["lon"])
                return GeocodeResult(lat=lat, lon=lon, source="nominatim", raw=data[0])
            except (KeyError, ValueError, TypeError):
                return GeocodeResult(lat=None, lon=None, source="nominatim", raw={"error": "parse_error"})
        return GeocodeResult(lat=None, lon=None, source="nominatim", raw={"note": "no_result"})

    def geocode_with_fallbacks(self, base_address: str, postal: Optional[str]) -> GeocodeResult:
        """正規化+複数段のクエリで命中率を高める。"""
        # 段階的に問い合わせ
        candidates: List[str] = []
        a0 = normalize_space(to_halfwidth_basic(base_address))
        a0 = unify_dashes(strip_parentheses(a0))
        a1 = strip_building_details(a0)

        if postal:
            candidates.append(f"{a0} {postal}")
        candidates.append(a0)
        candidates.append("日本 " + a0)
        if a1 and a1 != a0:
            candidates.append(a1)
            candidates.append("日本 " + a1)

        # 広域サーチ（丁目/番地等を削る）
        a2 = simplify_block_level(a1 or a0)
        if a2 and a2 not in candidates:
            candidates.append(a2)

        # さらに短縮（先頭2語）
        toks = a2.split() if a2 else a0.split()
        if len(toks) >= 2:
            short2 = " ".join(toks[:2])
            if short2 not in candidates:
                candidates.append(short2)

        # 問い合わせ
        for q in candidates:
            try:
                res = self.geocode_once(q)
                if res.lat is not None and res.lon is not None:
                    return res
            except requests.HTTPError as e:
                code = e.response.status_code if e.response is not None else -1
                time.sleep(5 if code == 429 else 2)
            except requests.RequestException:
                time.sleep(2)

        return GeocodeResult(lat=None, lon=None, source="nominatim", raw={"note": "all_fallbacks_failed"})


class AddressCache:
    """住所→(lat, lon, source, raw) のキャッシュ（CSV保管）。成功のみ保存。"""

    def __init__(self, cache_csv: str):
        self.cache_csv = cache_csv
        self.map: Dict[str, GeocodeResult] = {}
        self._load()

    def _load(self):
        if not os.path.exists(self.cache_csv):
            return
        with open(self.cache_csv, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                addr = row.get("address", "").strip()
                if not addr:
                    continue
                lat_str = row.get("lat", "").strip()
                lon_str = row.get("lon", "").strip()
                if not lat_str or not lon_str:
                    # 失敗行は読み込まない
                    continue
                try:
                    lat = float(lat_str)
                    lon = float(lon_str)
                except ValueError:
                    continue
                source = row.get("source", "nominatim")
                raw = {}
                if row.get("raw_json"):
                    try:
                        raw = json.loads(row["raw_json"])
                    except json.JSONDecodeError:
                        raw = {"_raw_json_parse_error": row.get("raw_json")}
                self.map[addr] = GeocodeResult(lat=lat, lon=lon, source=source, raw=raw)

    def save(self):
        tmp = self.cache_csv + ".tmp"
        with open(tmp, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["address", "lat", "lon", "source", "raw_json"])
            writer.writeheader()
            for addr, res in self.map.items():
                if res.lat is None or res.lon is None:
                    continue  # 失敗は保存しない
                writer.writerow({
                    "address": addr,
                    "lat": res.lat,
                    "lon": res.lon,
                    "source": res.source,
                    "raw_json": json.dumps(res.raw, ensure_ascii=False),
                })
        os.replace(tmp, self.cache_csv)

    def get(self, address: str) -> Optional[GeocodeResult]:
        return self.map.get(address)

    def put(self, address: str, result: GeocodeResult):
        if result.lat is None or result.lon is None:
            return  # 失敗は保存しない
        self.map[address] = result


# -------------------------------
# DF処理
# -------------------------------

LAT_NAME_CANDIDATES = [LAT_COL, "lat", "latitude", "緯度(度)"]
LON_NAME_CANDIDATES = [LON_COL, "lon", "lng", "longitude", "経度(度)"]


def find_existing_geo_cols(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    """緯度・経度の既存カラム名を異名考慮で検出。"""
    lat_col = None
    lon_col = None
    for c in df.columns:
        if c in LAT_NAME_CANDIDATES or c.lower() in [x.lower() for x in LAT_NAME_CANDIDATES]:
            lat_col = c
        if c in LON_NAME_CANDIDATES or c.lower() in [x.lower() for x in LON_NAME_CANDIDATES]:
            lon_col = c
    return lat_col, lon_col


def geocode_dataframe(df: pd.DataFrame,
                      client: NominatimClient,
                      cache: AddressCache,
                      address_col: Optional[str] = None,
                      component_groups: Optional[List[List[str]]] = None,
                      postal_cols: Optional[List[str]] = None) -> pd.DataFrame:
    """DataFrame に緯度経度列を付加して返す。空欄のみ補完。"""
    component_groups = component_groups or DEFAULT_COMPONENT_COL_GROUPS
    postal_cols = postal_cols or DEFAULT_POSTAL_COLS

    addr_series = prepare_address_series(df, address_col, component_groups)

    # 既存の緯度/経度列を検出
    lat_existing_name, lon_existing_name = find_existing_geo_cols(df)
    lat_existing = df[lat_existing_name] if lat_existing_name in df.columns else pd.Series([None] * len(df), index=df.index)
    lon_existing = df[lon_existing_name] if lon_existing_name in df.columns else pd.Series([None] * len(df), index=df.index)

    out = df.copy()
    if LAT_COL not in out.columns:
        out[LAT_COL] = None
    if LON_COL not in out.columns:
        out[LON_COL] = None

    for idx in tqdm(range(len(df)), desc="Geocoding"):
        # 既存に値があれば尊重
        lat0 = None
        lon0 = None
        if lat_existing_name and pd.notna(lat_existing.iloc[idx]):
            try:
                lat0 = float(lat_existing.iloc[idx])
            except Exception:
                lat0 = None
        if lon_existing_name and pd.notna(lon_existing.iloc[idx]):
            try:
                lon0 = float(lon_existing.iloc[idx])
            except Exception:
                lon0 = None
        if lat0 is not None and lon0 is not None:
            out.at[idx, LAT_COL] = lat0
            out.at[idx, LON_COL] = lon0
            continue

        # 住所と郵便番号
        addr = addr_series.iloc[idx]
        addr_n = normalize_space(to_halfwidth_basic(addr))
        if not addr_n:
            continue

        postal = detect_postal_code(df.iloc[idx], postal_cols)

        # キャッシュ確認（正規化後のアドレスキー）
        cached = cache.get(addr_n if not postal else f"{addr_n} {postal}")
        if cached and cached.lat is not None and cached.lon is not None:
            out.at[idx, LAT_COL] = cached.lat
            out.at[idx, LON_COL] = cached.lon
            continue

        # フォールバック付きジオコーディング
        res = client.geocode_with_fallbacks(addr_n, postal)
        if res.lat is not None and res.lon is not None:
            out.at[idx, LAT_COL] = res.lat
            out.at[idx, LON_COL] = res.lon
            cache.put(addr_n if not postal else f"{addr_n} {postal}", res)

    return out


# -------------------------------
# 保存（keep_vba 安定対応）
# -------------------------------

def save_excel_preserving_macros(input_path: str,
                                 output_path: str,
                                 frames_by_sheet: Dict[str, pd.DataFrame],
                                 keep_vba: bool):
    """
    ExcelにDataFrameを書き戻す。
    - keep_vba=True の場合、.xlsm のマクロを保持して保存
      → openpyxl.load_workbook(keep_vba=True) で既存ブックを開き、
         pandas.ExcelWriter(engine='openpyxl') に book として渡す。
      ※ pandas / openpyxl のバージョン差異による keep_vba エラーを回避。
    - keep_vba=False の場合、通常の .xlsx として保存（既存装飾は保持されません）。
    既存シートは DataFrame で完全置換します。
    """
    if keep_vba:
        # 入力ブックをマクロ保持で開く
        wb = load_workbook(input_path, keep_vba=True)
        # 既存同名シートを削除してから同名で再作成し、pandasで書き込む
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            writer.book = wb
            writer.sheets = {ws.title: ws for ws in wb.worksheets}

            for sheet_name, df in frames_by_sheet.items():
                if sheet_name in writer.sheets:
                    # 位置を維持するため、同じインデックスに作り直す
                    idx = wb.sheetnames.index(sheet_name)
                    wb.remove(wb[sheet_name])
                    wb.create_sheet(title=sheet_name, index=idx)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            writer.save()
    else:
        # 通常のxlsxとして全シート書き出し（完全置換）
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
            for sheet_name, df in frames_by_sheet.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)


# -------------------------------
# ワークブック処理
# -------------------------------

def process_workbook(input_path: str,
                     output_path: Optional[str],
                     sheets: Optional[List[str]] = None,
                     address_cols: Optional[List[str]] = None,
                     cache_csv: Optional[str] = None,
                     keep_vba: bool = False,
                     email: Optional[str] = None):
    """ブック全体を処理し、指定シートに緯度・経度を付与して保存。"""
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_path}")

    addr_candidates = address_cols or DEFAULT_ADDRESS_CANDIDATE_COLS

    # シート一覧
    xls = pd.ExcelFile(input_path, engine="openpyxl")
    all_sheets = xls.sheet_names
    target_sheets = sheets or all_sheets  # 指定がなければ全シート

    # キャッシュ
    cache_csv = cache_csv or os.path.splitext(input_path)[0] + "_geocode_cache.csv"
    cache = AddressCache(cache_csv)

    client = NominatimClient(email=email)

    frames_out: Dict[str, pd.DataFrame] = {}
    for s in target_sheets:
        df = pd.read_excel(input_path, sheet_name=s, engine="openpyxl")
        addr_col = detect_address_column(df, addr_candidates)
        df2 = geocode_dataframe(df, client, cache, address_col=addr_col)
        frames_out[s] = df2

    cache.save()

    # 出力パス決定
    if output_path is None:
        root, ext = os.path.splitext(input_path)
        if keep_vba and ext.lower() == ".xlsm":
            output_path = f"{root}_geocoded.xlsm"
        else:
            output_path = f"{root}_geocoded.xlsx"

    save_excel_preserving_macros(input_path, output_path, frames_out, keep_vba=keep_vba)

    return output_path, cache_csv, list(frames_out.keys())


# -------------------------------
# CLI
# -------------------------------

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Excelの住所列から緯度・経度を付与して保存します（Nominatim使用・日本語最適化・マクロ保持対応）。"
    )
    p.add_argument("input", help="入力Excelファイル（.xlsx / .xlsm）")
    p.add_argument("--output", help="出力Excelファイル（未指定時は *_geocoded.xlsx など自動命名）")
    p.add_argument("--sheets", nargs="*", help="処理対象のシート名（未指定なら全シート）")
    p.add_argument("--address-cols", nargs="*", help="住所候補カラム名（優先順）。未指定時は既定を使用")
    p.add_argument("--cache-csv", help="住所キャッシュCSVのパス。未指定時は <入力名>_geocode_cache.csv")
    p.add_argument("--keep-vba", action="store_true", help=".xlsm のマクロを保持して保存（環境差異でも安定）")
    p.add_argument("--email", help="Nominatim User-Agent 用の連絡先メール（.env の GEO_EMAIL でも可）")
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    try:
        out_path, cache_path, sheets = process_workbook(
            input_path=args.input,
            output_path=args.output,
            sheets=args.sheets,
            address_cols=args.address_cols,
            cache_csv=args.cache_csv,
            keep_vba=args.keep_vba,
            email=args.email,
        )
        print("✅ 完了")
        print(f"出力ファイル: {out_path}")
        print(f"キャッシュ   : {cache_path}")
        print(f"処理シート   : {', '.join(sheets)}")
        return 0
    except Exception as e:
        print("❌ エラー:", e, file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
